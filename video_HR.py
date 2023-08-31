import cv2
import numpy as np
import matplotlib.pyplot as plt
import mediapipe as mp
import heartpy as hp
from scipy import signal
import scipy.signal as sig
from SixDRepNet import SixDRepNet
from feat.utils import read_pictures
import math
import time
import os

model = SixDRepNet()
import openpyxl

dir = 'D:/Sekolah/111-1/Project/dataset'
directory = os.listdir(dir)

# Butterworth forward-backward band-pass filter
def bandpass(signal, fs, order, fc_low, fc_hig, debug=False):
    """Butterworth forward-backward band-pass filter.

    :param signal: list of ints or floats; The vector containing the signal samples.
    :param fs: float; The sampling frequency in Hz.
    :param order: int; The order of the filter.
    :param fc_low: int or float; The lower cutoff frequency of the filter.
    :param fc_hig: int or float; The upper cutoff frequency of the filter.
    :param debug: bool, default=False; Flag to enable the debug mode that prints additional information.

    :return: list of floats; The filtered signal.
    """
    nyq = 0.5 * fs  # Calculate the Nyquist frequency.
    cut_low = fc_low / nyq  # Calculate the lower cutoff frequency (-3 dB).
    cut_hig = fc_hig / nyq  # Calculate the upper cutoff frequency (-3 dB).
    bp_b, bp_a = sig.butter(order, (cut_low, cut_hig), btype="bandpass")  # Design and apply the band-pass filter.
    bp_data = list(sig.filtfilt(bp_b, bp_a, signal))  # Apply forward-backward filter with linear phase.
    return bp_data


# Fast Fourier Transform
def fft(data, fs, scale="mag"):
    # Apply Hanning window function to the data.
    data_win = data * np.hanning(len(data))
    if scale == "mag":  # Select magnitude scale.
        mag = 2.0 * np.abs(np.fft.rfft(tuple(data_win)) / len(data_win))  # Single-sided DFT -> FFT
    elif scale == "pwr":  # Select power scale.
        mag = np.abs(np.fft.rfft(tuple(data_win))) ** 2  # Spectral power
    bin = np.fft.rfftfreq(len(data_win), d=1.0 / fs)  # Calculate bins, single-sided
    return bin, mag


mp_drawing = mp.solutions.drawing_utils
mp_face_mesh = mp.solutions.face_mesh
wb_obj = openpyxl.load_workbook('data1.xlsx')

for folder in directory:
    files = os.listdir(dir+'/'+folder)
    for file in files:
        if file.endswith('.mp4'):
            # Excel parameters
            time_stamp = []
            blue = []
            red = []
            green = []

            # plotting parameters
            b_plot = []
            g_plot = []
            r_plot = []
            t_plot = []

            # Get source_mp4 Video
            source_mp4 = dir+'/'+ folder +'/'+file
            #workbook = Workbook(source_mp4[-9:]+'.xlsx')

            # Using Video-capture to get the fps value.
            capture = cv2.VideoCapture(source_mp4)
            fps = capture.get(cv2.CAP_PROP_FPS)
            vid_height = capture.get(cv2.CAP_PROP_FRAME_HEIGHT)
            vid_width = capture.get(cv2.CAP_PROP_FRAME_WIDTH)
            capture.release()

            # Using Video-capture to run video file
            cap = cv2.VideoCapture(source_mp4)

            frame_count = 0  # frames count
            time_count = 0  # time in milliseconds
            update = 0  # plot update
            plot = False  # True to show POS plots
            is_update = False
            pitch_arr = []
            yaw_arr = []
            roll_arr = []
            head_trackingX = []
            head_trackingY = []
            movement_length = []
            beforeX = 0
            beforeY = 0
            before_yaw = 0
            before_pitch = 0
            before_roll = 0
            yaw_change = []
            pitch_change = []
            roll_change = []
            blink_count = 0
            temp_check = 0

            def euclaideanDistance(point, point1):
                x, y = point
                x1, y1 = point1
                distance = math.sqrt((x1 - x)**2 + (y1 - y)**2)
                return distance

            # For webcam input:
            drawing_spec = mp_drawing.DrawingSpec(thickness=1, circle_radius=1)
            with mp_face_mesh.FaceMesh(min_detection_confidence=0.5, min_tracking_confidence=0.5, refine_landmarks=True) as face_mesh:
                while True:
                    success, image = cap.read()
                    if image is None:
                        break
                    height, width, _ = image.shape
                    # Flip the image horizontally for a later selfie-view display, and convert the BGR image to RGB.
                    image = cv2.cvtColor(cv2.flip(image, 1), cv2.COLOR_BGR2RGB)
                    # To improve performance, optionally mark the image as not writeable to pass by reference.
                    image.flags.writeable = False
                    processed_img = face_mesh.process(image)

                    # Draw the face mesh annotations on the image.
                    image.flags.writeable = True
                    image = cv2.cvtColor(image, cv2.COLOR_RGB2BGR)  # convert the RGB image to BGR.

                    pitch, yaw, roll = model.predict(image)

                    model.draw_axis(image, yaw, pitch, roll)
                    if before_pitch != 0 or before_roll != 0 or before_yaw != 0:
                        pitch_change.append(abs(pitch-before_pitch))
                        yaw_change.append(abs(yaw-before_yaw))
                        roll_change.append(abs(roll-before_roll))
                    pitch_arr.append(pitch)
                    yaw_arr.append(yaw)
                    roll_arr.append(roll)
                    before_yaw = yaw
                    before_pitch = pitch
                    before_roll = roll

                    if processed_img.multi_face_landmarks:
                        for face_landmarks in processed_img.multi_face_landmarks:
                            landmark_points = []
                            for i in range(0, 468):
                                x = int(face_landmarks.landmark[i].x * width)
                                y = int(face_landmarks.landmark[i].y * height)
                                p = [x, y]
                                landmark_points.append([x, y])
                            # Set ROI points
                            forehead = np.array((
                                landmark_points[9], landmark_points[107], landmark_points[66], landmark_points[105],
                                landmark_points[104], landmark_points[103],
                                landmark_points[67], landmark_points[109], landmark_points[10],
                                landmark_points[338], landmark_points[297], landmark_points[332],
                                landmark_points[333], landmark_points[334], landmark_points[296],
                                landmark_points[336]))
                            left_cheek = np.array((landmark_points[266], landmark_points[426], landmark_points[436],
                                                landmark_points[416], landmark_points[376],
                                                landmark_points[352], landmark_points[347], landmark_points[330]))
                            right_cheek = np.array((landmark_points[36], landmark_points[206], landmark_points[216],
                                                    landmark_points[192], landmark_points[147],
                                                    landmark_points[123], landmark_points[117], landmark_points[118],
                                                    landmark_points[101]))
                            left_eye = np.array((landmark_points[362], landmark_points[385], landmark_points[387], landmark_points[263], landmark_points[373], landmark_points[380]))
                            right_eye = np.array((landmark_points[33], landmark_points[160], landmark_points[158], landmark_points[133], landmark_points[153], landmark_points[144]))

                            # for tracking head (head movement and velocity)
                            if beforeX > 0 and beforeY > 0:
                                head_trackingX.append(landmark_points[8][0] - beforeX)
                                head_trackingY.append(landmark_points[8][1] - beforeY)
                                movement_length.append(euclaideanDistance([landmark_points[8][0],landmark_points[8][1]], [beforeX, beforeY]))
                            beforeX = landmark_points[8][0]
                            beforeY = landmark_points[8][1]

                            cv2.circle(image,(landmark_points[8][0],landmark_points[8][1]),3,(0,255,0),-1)

                            # mask the image and crop the ROI with black background
                            mask = np.zeros((height, width), dtype=np.uint8)
                            cv2.fillPoly(mask, [forehead, left_cheek, right_cheek], (255))
                            crop_img = cv2.bitwise_and(image, image, mask=mask)

                            # In case there are hair that get into the masked frame
                            # converting from BGR to hsv color space
                            img_HSV = cv2.cvtColor(crop_img, cv2.COLOR_BGR2HSV)
                            #skin color range for hsv color space 
                            HSV_mask = cv2.inRange(img_HSV, (0, 15, 0), (17,170,255)) 
                            HSV_mask = cv2.morphologyEx(HSV_mask, cv2.MORPH_OPEN, np.ones((3,3), np.uint8))

                            # converting from BGR to YCbCr color space
                            img_YCrCb = cv2.cvtColor(crop_img, cv2.COLOR_BGR2YCrCb)
                            #skin color range for hsv color space 
                            YCrCb_mask = cv2.inRange(img_YCrCb, (0, 135, 85), (255,180,135)) 
                            YCrCb_mask = cv2.morphologyEx(YCrCb_mask, cv2.MORPH_OPEN, np.ones((3,3), np.uint8))

                            # merge skin detection (YCbCr and hsv)
                            global_mask=cv2.bitwise_and(YCrCb_mask,HSV_mask)
                            global_mask=cv2.medianBlur(global_mask,3)
                            global_mask = cv2.morphologyEx(global_mask, cv2.MORPH_OPEN, np.ones((4,4), np.uint8))

                            final_image=cv2.bitwise_and(crop_img, crop_img, mask=global_mask)

                            # eliminate the black pixels and get mean of RGB for each frame
                            b, g, r = cv2.split(final_image)
                            indices_list = np.where(np.any(crop_img != [0, 0, 0], axis=-1))
                            roi_pixel_img = final_image[indices_list]

                            # Append the current frame's RGB to plotting parameters
                            b_plot.append(roi_pixel_img[:, 0].mean())
                            g_plot.append(roi_pixel_img[:, 1].mean())
                            r_plot.append(roi_pixel_img[:, 2].mean())
                            frame_count += 1
                            t_plot.append(round(time_count))
                            time_count += (1000 / fps)

                            # Draw the face mesh on the image
                            mp_drawing.draw_landmarks(
                                image=image,
                                landmark_list=face_landmarks,
                                connections=mp_face_mesh.FACEMESH_CONTOURS,
                                landmark_drawing_spec=drawing_spec,
                                connection_drawing_spec=drawing_spec)
                            cv2.imshow('MediaPipe FaceMesh', image)

                            
                            #Check the blinking rate
                                # Right eyes
                            temp1 = euclaideanDistance(right_eye[1], right_eye[5])
                            temp2 = euclaideanDistance(right_eye[2], right_eye[4])
                            temp3 = euclaideanDistance(right_eye[0], right_eye[3])
                            EAR_right = (temp1 + temp2)/(2*temp3)

                            #left eyes
                            temp1 = euclaideanDistance(left_eye[1], left_eye[5])
                            temp2 = euclaideanDistance(left_eye[2], left_eye[4])
                            temp3 = euclaideanDistance(left_eye[0], left_eye[3])
                            EAR_left = (temp1 + temp2)/(2*temp3)

                            EAR_avg = (EAR_right + EAR_left)/2

                            #blinking last for over than 100ms
                            if EAR_avg < 0.24 and temp_check == 0:
                                t1 = time.time()
                                temp_check = 1
                            elif EAR_avg >= 0.24 and temp_check == 1:
                                t2 = time.time()
                                time_length = t2-t1
                                temp_check = 0
                                if time_length >= 0.1:
                                    blink_count = blink_count + 1

                            # Plot the graph 4 times a sec (15 new records each time)
                            if frame_count % 15 == 0:
                                is_update = True  # New frame has come

                                update += 1

                            elif update > 2:
                                # After 3 plots push the reading to Excel parameters and clear plotting parameters
                                if is_update:
                                    if update == 3:
                                        blue.extend(b_plot)
                                        green.extend(g_plot)
                                        red.extend(r_plot)
                                        time_stamp.extend(t_plot)
                                    else:
                                        blue.extend(b_plot[(len(b_plot) - 15):len(b_plot)])
                                        green.extend(g_plot[(len(g_plot) - 15):len(g_plot)])
                                        red.extend(r_plot[(len(r_plot) - 15):len(r_plot)])
                                        time_stamp.extend(t_plot[(len(t_plot) - 15):len(t_plot)])

                                    del b_plot[0:15]
                                    del g_plot[0:15]
                                    del r_plot[0:15]
                                    del t_plot[0:15]

                                    is_update = False  # we added the new frame to our list structure

                    # Break using esc key
                    if cv2.waitKey(1) & 0xFF == 27:
                        break

                cv2.destroyAllWindows()
                cap.release()
                capture.release()


                # stack r, g, b channels into a single 2-D array
                mean_rgb = np.vstack((red, green, blue)).T

                # Calculating window length l and initiate bvp as 0's
                l = int(fps * 1.6)
                H = np.zeros(mean_rgb.shape[0])

                # POS Algorithm to extract bvp from raw signal
                for t in range(0, (mean_rgb.shape[0] - l)):
                    # Step 1: Spatial averaging
                    C = mean_rgb[t:t + l - 1, :].T

                    # Step 2 : Temporal normalization
                    mean_color = np.mean(C, axis=1)
                    diag_mean_color = np.diag(mean_color)
                    diag_mean_color_inv = np.linalg.inv(diag_mean_color)
                    Cn = np.matmul(diag_mean_color_inv, C)

                    # Step 3: projection_matrix
                    projection_matrix = np.array([[0, 1, -1], [-2, 1, 1]])
                    S = np.matmul(projection_matrix, Cn)
                        
                    # Step 4: 2D signal to 1D signal
                    std = np.array([1, np.std(S[0, :]) / np.std(S[1, :])])
                    # print("std", std)
                    P = np.matmul(std, S)

                    # Step 5: Overlap-Adding
                    H[t:t + l - 1] = H[t:t + l - 1] + (P - np.mean(P)) / np.std(P)

                # print("Pulse", H)
                bvp_signal = H

                # 2nd order butterworth bandpass filtering
                filtered_pulse = bandpass(bvp_signal, fps, 2, 0.9, 1.8)  # Heart Rate : 60-100 bpm (1-1.7 Hz), taking 54-108 (0.9 - 1.8)

                # plot welch's periodogram
                bvp_signal = bvp_signal.flatten()
                f_set, f_psd = signal.welch(bvp_signal, fps, window='hamming', nperseg=1024)  # , scaling='spectrum',nfft=2048)

                # Filtering the welch's periodogram - Heart Rate : 60-100 bpm (1-1.7 Hz), taking 54-108 (0.9 - 1.8)
                # green_psd = green_psd.flatten()
                first = np.where(f_set > 0.9)[0]  # 0.8 for 300 frames
                last = np.where(f_set < 1.8)[0]
                first_index = first[0]
                last_index = last[-1]
                range_of_interest = range(first_index, last_index + 1, 1)

                # get the frequency with highest psd
                # print("Range of interest", range_of_interest)
                max_idx = np.argmax(f_psd[range_of_interest])
                f_max = f_set[range_of_interest[max_idx]]

                # Calculate and display FFT of filtered pulse
                X_fft, Y_fft = fft(filtered_pulse, fps, scale="mag")

                # Welch's Periodogram of filtered pulse
                f_set, Pxx_den = signal.welch(filtered_pulse, fps, window='hamming', nperseg=1024)

                # Calculate Heart Rate and Plot using HeartPy Library
                working_data, measures = hp.process(filtered_pulse, fps)
                peaks = [0] * len(working_data['hr'])
                for p, q in zip(working_data['peaklist'], working_data['binary_peaklist']):
                    if q == 1:
                        peaks[p] = 1
                detected_peaks_data = {i: peaks.count(i) for i in peaks}
                #print('Detected number of peaks = ', detected_peaks_data[1])
                head_trackingX = np.absolute(head_trackingX)
                head_trackingY = np.absolute(head_trackingY)
                averageX = np.average(head_trackingX)
                averageY = np.average(head_trackingY)
                averageMove = np.average(movement_length)
                sdX = np.std(head_trackingX)
                sdY = np.std(head_trackingY)
                sdMove = np.std(head_trackingY)
                print(measures)

            if file[-8:-6] == 'T1':
                # Export Data to Excel file
                sheet = wb_obj['Relax']
            else:
                sheet = wb_obj['Stress']

            row = 1
            col = 1
                
            if sheet['A1'].value is None:
                sheet.cell(row, col).value = 'Red Mean'
                sheet.cell(row, col+1).value = 'Green Mean'
                sheet.cell(row, col+2).value = 'Blue Mean'
                sheet.cell(row, col+3).value = 'Yaw Change(Mean)'
                sheet.cell(row, col+4).value = 'Roll Change(Mean)'
                sheet.cell(row, col+5).value = 'Pitch Change(Mean)'
                sheet.cell(row, col + 6).value = 'Yaw Change(Std)'
                sheet.cell(row, col + 7).value = 'Roll Change(Std)'
                sheet.cell(row, col + 8).value = 'Pitch Change(Std)'
                sheet.cell(row, col + 9).value = 'Heart Rate'
                sheet.cell(row, col + 10).value = 'Heart Rate Variability'
                sheet.cell(row, col + 11).value = 'Blinking Rate(per Min)'
                sheet.cell(row, col + 12).value = 'Head Move(Mean)'
                sheet.cell(row, col + 13).value = 'Head Move X(Std)'
                sheet.cell(row, col + 14).value = 'Head Move Y(Std)'
                sheet.cell(row, col + 15).value = 'Head Move(Std)'
                
            row = sheet.max_row+1
            sheet.cell(row, col).value = np.mean(red)
            sheet.cell(row, col+1).value = np.mean(green)
            sheet.cell(row, col+2).value = np.mean(blue)
            sheet.cell(row, col+3).value = np.mean(yaw_change)
            sheet.cell(row, col+4).value = np.mean(roll_change)
            sheet.cell(row, col+5).value = np.mean(pitch_change)
            sheet.cell(row, col + 6).value = np.std(yaw_change)
            sheet.cell(row, col + 7).value = np.std(roll_change)
            sheet.cell(row, col + 8).value = np.std(pitch_change)
            sheet.cell(row, col + 9).value =  measures['bpm']
            sheet.cell(row, col + 10).value = measures['rmssd']
            sheet.cell(row, col + 11).value = blink_count*2
            sheet.cell(row, col + 12).value = averageMove
            sheet.cell(row, col + 13).value = sdX
            sheet.cell(row, col + 14).value = sdY
            sheet.cell(row, col + 15).value = sdMove
            wb_obj.save('data1.xlsx')